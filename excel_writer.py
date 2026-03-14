import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

TEMPLATE_PATH = Path(__file__).parent / "admin_template.xlsx"
_DEFAULT_OUTPUT = str(Path(__file__).resolve().parent / "output")

FIELD_KEYS = [
    "customer_part_no",
    "customer_product_name",
    "product_model",
    "product_name",
    "brand",
    "quantity",
    "remark_customer",
    "remark_supply_chain",
    "customer_project_no",
    "customer_material_no",
    "inventory_feature",
    "major_category",
    "minor_category",
    "supply_org",
    "attachment_filename",
    "remark_purchase",
    "customer_expected_delivery",
    "customer_expected_price",
    "warehouse_factory",
    "sales_unit_price_tax",
    "shipment_date",
]

_FS_INVALID_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]+')
_MULTI_UNDERSCORE = re.compile(r"_+")
_MAX_SEGMENT_CHARS = 50
_OPENCLAW_UUID_RE = re.compile(r"---[0-9a-f][-0-9a-f]*$", re.IGNORECASE)
_ORDER_NO_RE = re.compile(r"[A-Z]{2,5}\d{6,15}")


def _strip_openclaw_suffix(name: str) -> str:
    m = _OPENCLAW_UUID_RE.search(name)
    return name[:m.start()].rstrip("_") if m else name


def _extract_order_no(name: str) -> str:
    m = _ORDER_NO_RE.search(name)
    return m.group() if m else ""


def _sanitize_segment(value: str, fallback: str = "unknown") -> str:
    cleaned = _FS_INVALID_CHARS.sub("_", (value or "").strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = cleaned.rstrip(" .")
    cleaned = _MULTI_UNDERSCORE.sub("_", cleaned).strip("_")
    if not cleaned:
        cleaned = _MULTI_UNDERSCORE.sub("_", _FS_INVALID_CHARS.sub("_", fallback).strip("_. ")) or "unknown"
    return cleaned[:_MAX_SEGMENT_CHARS]


def _build_output_name(source_file: str, company_name: str = "") -> str:
    order_no = _extract_order_no(Path(source_file).stem)
    company = company_name.strip()
    if order_no and company:
        raw = f"{order_no}_{company}"
    elif order_no:
        raw = order_no
    elif company:
        raw = company
    else:
        raw = _strip_openclaw_suffix(Path(source_file).stem)
    stem = _sanitize_segment(raw, fallback="bom")
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"admin_template_{stem}_{ts}.xlsx"


def write_admin_template(
    rows: list[dict],
    source_file: str,
    output_dir: str = "",
    company_name: str = "",
) -> str:
    root = Path(output_dir).resolve() if output_dir else Path(_DEFAULT_OUTPUT)
    month = datetime.now().strftime("%Y-%m")
    fallback_stem = _strip_openclaw_suffix(Path(source_file).stem) or "unknown"
    company_seg = _sanitize_segment(company_name, fallback=fallback_stem)
    target_dir = root / month / company_seg
    target_dir.mkdir(parents=True, exist_ok=True)

    output_name = _build_output_name(source_file, company_name=company_name)
    output_path = (target_dir / output_name).resolve()
    shutil.copy2(str(TEMPLATE_PATH), str(output_path))

    wb = load_workbook(str(output_path))
    ws = wb.active

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, key in enumerate(FIELD_KEYS, start=1):
            value = row_data.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(str(output_path))
    wb.close()
    return str(output_path)

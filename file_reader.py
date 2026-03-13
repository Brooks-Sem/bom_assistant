import base64
import logging
from pathlib import Path

from openpyxl import load_workbook

try:
    import fitz
except ImportError:
    fitz = None

log = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".xlsx", ".csv", ".pdf", ".png", ".jpg", ".jpeg"}
_PDF_MIN_ALNUM_CHARS = 40
_PDF_MIN_LINES = 2
_PDF_MIN_ALNUM_RATIO = 0.15


def detect_file_type(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    if ext == ".xlsx":
        return "xlsx"
    if ext == ".csv":
        return "csv"
    if ext == ".pdf":
        return "pdf"
    if ext in {".png", ".jpg", ".jpeg"}:
        return "image"
    raise ValueError(f"不支持的文件格式: {ext}")


def read_xlsx_as_text(file_path: str) -> str:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parts.append(f"[Sheet: {sheet_name}]")
        headers = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("\t".join(headers))
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)


def read_csv_as_text(file_path: str) -> str:
    return Path(file_path).read_text(encoding="utf-8")


def _has_usable_text(text: str) -> bool:
    if not text:
        return False
    non_ws = [ch for ch in text if not ch.isspace()]
    if not non_ws:
        return False
    alnum = sum(c.isalnum() for c in non_ws)
    if alnum < _PDF_MIN_ALNUM_CHARS:
        return False
    if alnum / len(non_ws) < _PDF_MIN_ALNUM_RATIO:
        return False
    meaningful_lines = sum(
        1 for line in text.splitlines() if any(ch.isalnum() for ch in line)
    )
    return meaningful_lines >= _PDF_MIN_LINES


def _extract_table_lines(page) -> list[str]:
    if not hasattr(page, "find_tables"):
        return []
    try:
        finder = page.find_tables()
    except Exception:
        return []
    lines: list[str] = []
    for idx, table in enumerate(getattr(finder, "tables", None) or [], 1):
        try:
            rows = table.extract() or []
        except Exception:
            continue
        table_rows: list[str] = []
        for row in rows:
            if not row:
                continue
            cells = [" ".join(str(c or "").split()) for c in row]
            if any(cells):
                table_rows.append("\t".join(cells))
        if table_rows:
            lines.append(f"[Table {idx}]")
            lines.extend(table_rows)
    return lines


def read_pdf_as_text(file_path: str) -> str:
    if fitz is None:
        return ""
    try:
        doc = fitz.open(file_path)
    except Exception:
        log.warning("PyMuPDF 无法打开 PDF: %s", file_path, exc_info=True)
        return ""
    try:
        parts: list[str] = []
        for i in range(doc.page_count):
            try:
                page = doc.load_page(i)
                raw = page.get_text("text", sort=True).strip()
            except Exception:
                log.warning("PDF 第 %d 页文本提取失败，跳过", i + 1, exc_info=True)
                continue
            tables = _extract_table_lines(page)

            page_lines: list[str] = []
            if raw:
                page_lines.append(f"[Page {i + 1}]")
                page_lines.extend(
                    line.rstrip() for line in raw.splitlines() if line.strip()
                )
            if tables:
                if not page_lines:
                    page_lines.append(f"[Page {i + 1}]")
                page_lines.extend(tables)

            if page_lines:
                parts.append("\n".join(page_lines))

        text = "\n\n".join(parts).strip()
        if not _has_usable_text(text):
            log.info("PDF 文本质量不足，判定为扫描件: %s", file_path)
            return ""
        return text
    except Exception:
        log.warning("PDF 文本提取异常: %s", file_path, exc_info=True)
        return ""
    finally:
        doc.close()


def read_file_as_base64(file_path: str) -> str:
    data = Path(file_path).read_bytes()
    return base64.standard_b64encode(data).decode("ascii")


def get_pdf_media_type() -> str:
    return "application/pdf"


def get_image_media_type(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    mapping = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}
    return mapping.get(ext, "application/octet-stream")
